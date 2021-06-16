import tkinter as tk
from openpyxl import Workbook, load_workbook

class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.pack()
        self.labels = ["First Name:",
                       "Last Name:",
                       "Address Line 1:",
                       "Address Line 2:",
                       "City:",
                       "State/Province:",
                       "Postal Code:",
                       "Country:",
                      ]
    def run(self):
        self.open_workbook()              
        self.create_widgets()
        self.master.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.master.mainloop()

    def open_workbook(self):
        sheet_exist = False
        try:
            self.wb = load_workbook('addresses.xlsx')
            sheet_exist = True
        except FileNotFoundError:
            self.wb = Workbook()
        self.sheet = self.wb.active

        if sheet_exist:
            self.min_row = self.sheet.max_row + 1
        else:
            self.min_row = 1
            for header_row in self.sheet.iter_rows(min_row=1, max_col=8, max_row=1):
                self.min_row += 1
                idx = 0
                for cell in header_row:
                    cell.value = self.labels[idx].strip(":")
                    idx +=1

    def update_sheet(self):
        address = list()
        for entry_widget in self.entry_widgets:
            address.append(entry_widget.get())
        if not any(address):
            pass
        else:
            for row in self.sheet.iter_rows(min_row=self.min_row, max_col=8, max_row=self.min_row):
                self.min_row += 1
                idx = 0 
                for cell in row:
                    cell.value = address[idx]
                    idx += 1
                address = list()     

    def clear_entries(self):
        for entry_widget in self.entry_widgets:
            entry_widget.delete(0, tk.END)

    def on_closing(self):
        self.wb.save('addresses.xlsx')
        self.master.destroy()
                
    def create_widgets(self):
        frm_form = tk.Frame(master=self, relief=tk.SUNKEN, borderwidth=3)
        # Pack the frame into the window
        frm_form.pack()

        self.entry_widgets = list()

        # Loop over the list of field labels and create label and entry widgets.
        # Also store the entry widgets in a list
        for idx, text in enumerate(self.labels):
            # Create a Label widget with the text from the labels list
            label = tk.Label(master=frm_form, text=text)
            entry = tk.Entry(master=frm_form, width=50)

            self.entry_widgets.append(entry)
            
            # Use the grid geometry manager to position the widgets
            label.grid(row=idx, column=0, sticky="e")
            entry.grid(row=idx, column=1)

        # Create buttons frame
        frm_buttons = tk.Frame(master=self)
        frm_buttons.pack(fill=tk.X, ipadx=5, ipady=5)

        # Create the "Submit" button
        btn_submit = tk.Button(master=frm_buttons, text="UpdateXL", command=self.update_sheet)
        btn_submit.pack(side=tk.RIGHT, padx=10, ipadx=10)
        
        #Create the "Clear" button
        btn_clear = tk.Button(master=frm_buttons, text="Clear", command=self.clear_entries)
        btn_clear.pack(side=tk.RIGHT, ipadx=10)

        
if __name__ == "__main__":
    window = tk.Tk()
    window.title("Address Entry Form")
    app = Application(master=window)
    app.run()