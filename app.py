import tkinter as tk
from openpyxl import Workbook, load_workbook

# List of field labels
labels = [
    "First Name:",
    "Last Name:",
    "Address Line 1:",
    "Address Line 2:",
    "City:",
    "State/Province:",
    "Postal Code:",
    "Country:",
]

sheet_exist = False
try:
    wb = load_workbook('addresses.xlsx')
    sheet_exist = True
except FileNotFoundError:
    wb = Workbook()
sheet = wb.active

if sheet_exist:
    min_row = sheet.max_row + 1
else:
    min_row = 1
    for header_row in sheet.iter_rows(min_row=1, max_col=8, max_row=1):
        min_row += 1
        idx = 0
        for cell in header_row:
            cell.value = labels[idx].strip(":")
            idx +=1

def update_sheet():
    global min_row
    address = list()
    for entry_widget in entry_widgets:
        address.append(entry_widget.get())
    if not any(address):
        pass
    else:
        for row in sheet.iter_rows(min_row=min_row, max_col=8, max_row=min_row):
            min_row += 1
            idx = 0 
            for cell in row:
                cell.value = address[idx]
                idx += 1
            address = list()     

def clear_entries():
    for entry_widget in entry_widgets:
        entry_widget.delete(0, tk.END)

def on_closing():
    wb.save('addresses.xlsx')
    window.destroy()

# Create a new window with the title "Address Entry Form"
window = tk.Tk()
window.title("Address Entry Form")

# Create a new frame `frm_form` to contain the Label
# and Entry widgets for entering address information.
frm_form = tk.Frame(relief=tk.SUNKEN, borderwidth=3)
# Pack the frame into the window
frm_form.pack()

entry_widgets = list()
# Loop over the list of field labels
for idx, text in enumerate(labels):
    # Create a Label widget with the text from the labels list
    label = tk.Label(master=frm_form, text=text)
    # Create an Entry widget
    entry = tk.Entry(master=frm_form, width=50)
    entry_widgets.append(entry)
    # Use the grid geometry manager to place the Label and
    # Entry widgets in the row whose index is idx
    label.grid(row=idx, column=0, sticky="e")
    entry.grid(row=idx, column=1)

# Create a new frame `frm_buttons to contain the
# Submit and Clear buttons.
frm_buttons = tk.Frame()
frm_buttons.pack(fill=tk.X, ipadx=5, ipady=5)

# Create the "Submit" button
btn_submit = tk.Button(master=frm_buttons, text="UpdateXL", command=update_sheet)
btn_submit.pack(side=tk.RIGHT, padx=10, ipadx=10)

# Create the "Clear" button and pack it to the
# right side of `frm_buttons`
btn_clear = tk.Button(master=frm_buttons, text="Clear", command=clear_entries)
btn_clear.pack(side=tk.RIGHT, ipadx=10)

# Start the application
window.protocol("WM_DELETE_WINDOW", on_closing)
window.mainloop()